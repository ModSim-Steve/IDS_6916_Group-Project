import openpyxl
import csv
from typing import Dict, Tuple
import os
from WarGamingEnvironment_vTest import TerrainType, ElevationType


def excel_to_csv(excel_file_path: str, csv_file_path: str, max_width: int = 400, max_height: int = 100):
    # Check if the Excel file exists
    if not os.path.exists(excel_file_path):
        raise FileNotFoundError(f"Excel file not found: {excel_file_path}")

    # Load the Excel file
    wb = openpyxl.load_workbook(excel_file_path)
    sheet = wb.active

    # Define mappings using environment enums
    terrain_map = {
        'B': TerrainType.BARE.name,
        'S': TerrainType.SPARSE_VEG.name,
        'D': TerrainType.DENSE_VEG.name,
        'W': TerrainType.WOODS.name,
        'T': TerrainType.STRUCTURE.name
    }
    elevation_map = {
        'G': ElevationType.GROUND_LEVEL.name,
        'E': ElevationType.ELEVATED_LEVEL.name,
        'L': ElevationType.LOWER_LEVEL.name
    }

    # Check if the Excel sheet dimensions match the environment size
    if sheet.max_column > max_width or sheet.max_row > max_height:
        print(f"Warning: Excel sheet dimensions ({sheet.max_column}x{sheet.max_row}) "
              f"exceed the environment size ({max_width}x{max_height}).")
        print("Extra cells will be ignored.")
    elif sheet.max_column < max_width or sheet.max_row < max_height:
        print(f"Warning: Excel sheet dimensions ({sheet.max_column}x{sheet.max_row}) "
              f"are smaller than the environment size ({max_width}x{max_height}).")
        print("Missing cells will be filled with default values (BARE, GROUND_LEVEL).")

    # Ensure the directory for the CSV file exists
    os.makedirs(os.path.dirname(csv_file_path), exist_ok=True)

    # Write CSV with environment properties
    with open(csv_file_path, 'w', newline='') as f:
        writer = csv.writer(f)

        # Updated header to include environment properties
        writer.writerow(
            ['x', 'y', 'terrain_type', 'elevation_type', 'movement_cost', 'visibility_factor', 'cover_bonus'])

        # Iterate through cells
        for row in range(1, max_height + 1):
            for col in range(1, max_width + 1):
                # Get cell value or use default
                if row <= sheet.max_row and col <= sheet.max_column:
                    cell_value = sheet.cell(row=row, column=col).value
                else:
                    cell_value = None

                # Process terrain and elevation
                if cell_value:
                    terrain = terrain_map.get(cell_value[0], TerrainType.BARE.name)
                    elevation = elevation_map.get(cell_value[1], ElevationType.GROUND_LEVEL.name)
                else:
                    terrain = TerrainType.BARE.name
                    elevation = ElevationType.GROUND_LEVEL.name

                # Calculate environment properties based on terrain
                terrain_type = TerrainType[terrain]
                movement_cost = {
                    TerrainType.BARE: 1.0,
                    TerrainType.SPARSE_VEG: 1.2,
                    TerrainType.DENSE_VEG: 1.5,
                    TerrainType.WOODS: 2.0,
                    TerrainType.STRUCTURE: 3.0
                }[terrain_type]

                visibility_factor = {
                    TerrainType.BARE: 1.0,
                    TerrainType.SPARSE_VEG: 0.8,
                    TerrainType.DENSE_VEG: 0.5,
                    TerrainType.WOODS: 0.2,
                    TerrainType.STRUCTURE: 0.0
                }[terrain_type]

                cover_bonus = {
                    TerrainType.BARE: 0.0,
                    TerrainType.SPARSE_VEG: 0.1,
                    TerrainType.DENSE_VEG: 0.3,
                    TerrainType.WOODS: 0.6,
                    TerrainType.STRUCTURE: 0.9
                }[terrain_type]

                # Write row with all properties
                writer.writerow([
                    col - 1,  # x
                    row - 1,  # y
                    terrain,
                    elevation,
                    movement_cost,
                    visibility_factor,
                    cover_bonus
                ])

    print(f"CSV file '{csv_file_path}' has been created with dimensions {max_width}x{max_height}.")


if __name__ == "__main__":
    # Get the current script's directory
    script_dir = os.path.dirname(os.path.abspath(__file__))

    # Construct full paths for input and output files
    excel_file = os.path.join(script_dir, "map_design.xlsx")
    csv_file = os.path.join(script_dir, "generated_map.csv")

    # Check if the Excel file exists
    if not os.path.exists(excel_file):
        print(f"Excel file not found: {excel_file}")
        print("Please ensure the Excel file is in the same directory as this script.")
    else:
        print(f"Converting {excel_file} to CSV...")
        excel_to_csv(excel_file, csv_file)
